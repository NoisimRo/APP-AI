#!/usr/bin/env python3
"""Import CPV codes from an Excel (.xlsx) file into the database.

Reads the standard CPV nomenclator Excel file and populates the nomenclator_cpv table.
Also enriches existing decisions with CPV descriptions from the nomenclator.

Expected Excel structure (sheet: "CPV codes"):
  A: CODE (XXXXXXXX-X)
  B: denumire_ro
  C: denumire_en
  D: Categorie (Furnizare|Servicii|Lucrari)
  E: Clasa produse standardizate

Usage:
    # Import from local file
    DATABASE_URL="..." python scripts/import_cpv_xlsx.py --file Lista-coduri-CPV.xlsx

    # Import from GCS
    DATABASE_URL="..." python scripts/import_cpv_xlsx.py --gcs gs://date-expert-app/Lista-coduri-CPV.xlsx

    # Dry run (parse only, no DB writes)
    DATABASE_URL="..." python scripts/import_cpv_xlsx.py --file Lista-coduri-CPV.xlsx --dry-run

    # Also enrich decisions with CPV descriptions after import
    DATABASE_URL="..." python scripts/import_cpv_xlsx.py --file Lista-coduri-CPV.xlsx --enrich
"""

import asyncio
import argparse
import re
import sys
import tempfile
from pathlib import Path
from typing import Optional

# Add backend to path
sys.path.insert(0, str(Path(__file__).parent.parent / "backend"))

from sqlalchemy import select, delete, update, func
from app.core.logging import get_logger
from app.db.session import init_db
from app.db import session as db_session
from app.models.decision import NomenclatorCPV, DecizieCNSC

logger = get_logger(__name__)


def compute_cpv_level(cod_cpv: str) -> int:
    """Compute hierarchical level of a CPV code.

    Level 1: XX000000-X (Diviziune)
    Level 2: XXX00000-X (Grup)
    Level 3: XXXX0000-X (Clasă)
    Level 4: XXXXX000-X (Categorie)
    Level 5: XXXXXX00-X to XXXXXXXX-X (Subcategorie)
    """
    digits = cod_cpv[:8]
    level = 1
    for i in range(2, 8):
        if digits[i] != '0':
            level = i
    # Normalize levels > 5 to 5 (subcategorie)
    return min(level, 5)


def compute_cpv_parent(cod_cpv: str, all_codes: dict[str, str]) -> Optional[str]:
    """Compute parent CPV code by zeroing rightmost non-zero digits.

    Looks up the actual parent code (with correct check digit) from all_codes dict.
    all_codes maps first-8-digits → full code (e.g., '03111100' → '03111100-3').

    Args:
        cod_cpv: CPV code like "03111100-3".
        all_codes: Dict mapping 8-digit prefix to full code.

    Returns:
        Full parent code with correct check digit, or None for top-level.
    """
    digits = cod_cpv[:8]
    # Zero out from rightmost non-zero digit
    for i in range(7, 1, -1):
        if digits[i] != '0':
            parent_prefix = digits[:i] + '0' * (8 - i)
            if parent_prefix in all_codes:
                return all_codes[parent_prefix]
            # Keep zeroing if parent doesn't exist at this level
    return None


def parse_xlsx(filepath: str) -> list[dict]:
    """Parse the CPV Excel file.

    Args:
        filepath: Path to .xlsx file.

    Returns:
        List of CPV record dicts.
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Find the sheet with CPV data
    sheet = None
    for name in wb.sheetnames:
        if 'cpv' in name.lower() or 'code' in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    print(f"Using sheet: '{sheet.title}'")

    records = []
    code_re = re.compile(r'^\d{8}-\d$')

    for i, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), 1):
        if not row or not row[0]:
            continue

        code = str(row[0]).strip()
        if not code_re.match(code):
            # Skip header rows
            if i <= 3:
                continue
            # Log unexpected format
            if code and not code.upper().startswith('CODE'):
                logger.warning("skipping_invalid_code", row=i, code=code[:20])
            continue

        denumire_ro = str(row[1]).strip() if row[1] else ''
        denumire_en = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        categorie = str(row[3]).strip() if len(row) > 3 and row[3] else None
        clasa = str(row[4]).strip() if len(row) > 4 and row[4] else None

        # Normalize categorie
        if categorie:
            cat_lower = categorie.lower()
            if cat_lower.startswith('lucr'):
                categorie = 'Lucrări'
            elif cat_lower.startswith('serv'):
                categorie = 'Servicii'
            elif cat_lower.startswith('furn'):
                categorie = 'Furnizare'

        records.append({
            'cod_cpv': code,
            'descriere': denumire_ro,
            'denumire_en': denumire_en,
            'categorie_achizitii': categorie,
            'clasa_produse': clasa,
        })

    wb.close()
    return records


async def import_records(records: list[dict], force: bool = True) -> dict:
    """Import CPV records into the database.

    Args:
        records: Parsed CPV records.
        force: Truncate table before importing (default True for full replace).

    Returns:
        Import statistics.
    """
    stats = {
        'total_parsed': len(records),
        'imported': 0,
        'skipped': 0,
        'failed': 0,
        'errors': [],
    }

    # Build lookup for parent computation: 8-digit prefix → full code
    all_codes = {rec['cod_cpv'][:8]: rec['cod_cpv'] for rec in records}

    async with db_session.async_session_factory() as session:
        if force:
            await session.execute(delete(NomenclatorCPV))
            await session.commit()
            print("Deleted all existing CPV records.")

        # Batch insert for performance
        batch = []
        batch_size = 500

        for rec in records:
            try:
                parent = compute_cpv_parent(rec['cod_cpv'], all_codes)
                nivel = compute_cpv_level(rec['cod_cpv'])

                cpv = NomenclatorCPV(
                    cod_cpv=rec['cod_cpv'],
                    descriere=rec['descriere'],
                    denumire_en=rec.get('denumire_en'),
                    categorie_achizitii=rec['categorie_achizitii'],
                    clasa_produse=rec['clasa_produse'],
                    cod_parinte=parent,
                    nivel=nivel,
                )
                batch.append(cpv)
                stats['imported'] += 1

                if len(batch) >= batch_size:
                    session.add_all(batch)
                    await session.flush()
                    batch = []

            except Exception as e:
                stats['failed'] += 1
                stats['errors'].append(f"{rec['cod_cpv']}: {e}")
                logger.error("cpv_record_failed", cod=rec['cod_cpv'], error=str(e))

        # Flush remaining
        if batch:
            session.add_all(batch)
            await session.flush()

        await session.commit()
        print(f"Imported {stats['imported']} CPV codes.")

    return stats


async def enrich_decisions() -> dict:
    """Enrich decisions with CPV descriptions from nomenclator.

    Updates cpv_descriere, cpv_categorie, cpv_clasa on decizii_cnsc
    by matching cod_cpv with nomenclator_cpv.

    Returns:
        Enrichment statistics.
    """
    stats = {'updated': 0, 'no_match': 0, 'total': 0}

    async with db_session.async_session_factory() as session:
        # Load nomenclator into memory (small table)
        result = await session.execute(select(NomenclatorCPV))
        cpv_map = {
            cpv.cod_cpv: {
                'descriere': cpv.descriere,
                'categorie': cpv.categorie_achizitii,
                'clasa': cpv.clasa_produse,
            }
            for cpv in result.scalars().all()
        }
        # Also build prefix map (first 8 digits) for fuzzy matching
        prefix_map = {code[:8]: code for code in cpv_map}

        print(f"Loaded {len(cpv_map)} CPV codes for enrichment.")

        # Get all decisions with CPV codes
        result = await session.execute(
            select(DecizieCNSC.id, DecizieCNSC.cod_cpv)
            .where(DecizieCNSC.cod_cpv.isnot(None))
        )
        decisions = result.all()
        stats['total'] = len(decisions)

        for dec_id, cod_cpv in decisions:
            info = cpv_map.get(cod_cpv)

            # Try prefix match if exact match fails (check digit mismatch)
            if not info and cod_cpv and len(cod_cpv) >= 8:
                prefix = cod_cpv[:8]
                matched_code = prefix_map.get(prefix)
                if matched_code:
                    info = cpv_map[matched_code]

            if info:
                await session.execute(
                    update(DecizieCNSC)
                    .where(DecizieCNSC.id == dec_id)
                    .values(
                        cpv_descriere=info['descriere'],
                        cpv_categorie=info['categorie'],
                        cpv_clasa=info['clasa'],
                    )
                )
                stats['updated'] += 1
            else:
                stats['no_match'] += 1

        await session.commit()
        print(f"Enriched {stats['updated']} decisions, {stats['no_match']} without CPV match.")

    return stats


async def main():
    parser = argparse.ArgumentParser(
        description="Import CPV codes from Excel (.xlsx) into the database"
    )
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument(
        "--file", type=str,
        help="Path to local .xlsx file",
    )
    source.add_argument(
        "--gcs", type=str,
        help="GCS URI (e.g., gs://date-expert-app/Lista-coduri-CPV.xlsx)",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Parse only, print sample without writing to DB",
    )
    parser.add_argument(
        "--enrich", action="store_true",
        help="Also enrich existing decisions with CPV descriptions after import",
    )
    parser.add_argument(
        "--enrich-only", action="store_true",
        help="Skip import, only enrich decisions from existing nomenclator",
    )
    args = parser.parse_args()

    # Initialize database
    await init_db()

    if args.enrich_only:
        stats = await enrich_decisions()
        print(f"\nEnrichment complete: {stats}")
        return

    # Get file path
    filepath = args.file
    if args.gcs:
        # Download from GCS
        from google.cloud import storage as gcs_storage
        match = re.match(r'gs://([^/]+)/(.+)', args.gcs)
        if not match:
            print(f"ERROR: Invalid GCS URI: {args.gcs}")
            sys.exit(1)
        bucket_name, blob_path = match.groups()
        print(f"Downloading from GCS: {args.gcs}...")
        client = gcs_storage.Client()
        bucket = client.bucket(bucket_name)
        blob = bucket.blob(blob_path)
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        blob.download_to_filename(tmp.name)
        filepath = tmp.name
        print(f"Downloaded to {filepath}")

    # Parse
    print(f"Parsing {filepath}...")
    records = parse_xlsx(filepath)
    print(f"Parsed {len(records)} CPV codes")

    if not records:
        print("ERROR: No CPV records found. Check the file format.")
        sys.exit(1)

    # Show sample
    if args.dry_run or True:
        print(f"\nSample (first 15 + last 5):")
        print(f"{'Code':<14} {'Cat':<12} {'Nivel':<6} {'Descriere RO':<55} {'Clasa'}")
        print("-" * 140)
        all_codes = {r['cod_cpv'][:8]: r['cod_cpv'] for r in records}
        sample = records[:15] + records[-5:]
        for rec in sample:
            nivel = compute_cpv_level(rec['cod_cpv'])
            desc = rec['descriere'][:53]
            clasa = (rec['clasa_produse'] or '-')[:30]
            cat = (rec['categorie_achizitii'] or '-')[:10]
            print(f"{rec['cod_cpv']:<14} {cat:<12} {nivel:<6} {desc:<55} {clasa}")

        # Stats
        cats = {}
        levels = {}
        empty_desc = 0
        for rec in records:
            cat = rec['categorie_achizitii'] or 'N/A'
            cats[cat] = cats.get(cat, 0) + 1
            lvl = compute_cpv_level(rec['cod_cpv'])
            levels[lvl] = levels.get(lvl, 0) + 1
            if not rec['descriere']:
                empty_desc += 1

        print(f"\nPer categorie: {cats}")
        print(f"Per nivel: {dict(sorted(levels.items()))}")
        print(f"Descrieri goale: {empty_desc}")

    if args.dry_run:
        print("\n[DRY RUN] No database changes made.")
        return

    # Import
    print(f"\nImporting {len(records)} CPV codes (truncate + insert)...")
    stats = await import_records(records, force=True)
    print(f"\nImport stats: {stats}")

    if stats['errors']:
        print(f"\nFirst 10 errors:")
        for err in stats['errors'][:10]:
            print(f"  {err}")

    # Enrich decisions
    if args.enrich:
        print(f"\nEnriching decisions with CPV descriptions...")
        enrich_stats = await enrich_decisions()
        print(f"Enrichment stats: {enrich_stats}")

    print("\nDone!")


if __name__ == "__main__":
    asyncio.run(main())
